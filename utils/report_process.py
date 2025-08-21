
import os
import argparse

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import colors


class GcuOPInfo:
    def __init__(self, model_name, op_id, index, gcu_op, topsflame_op,
                 attribute, input_shape, output_shape, input_size, output_size, total_size,
                 floats_1d, floats_2d, busy, target, gap_to_pre, flame_start, top_flame_start, topsflame_duration,
                 launch_kernel_gap, top_topsflameop_handle, top_topsflameop_handle_name, start_timestamp, end_timestamp):
        self.model_name = model_name
        self.op_id = op_id
        self.index = index
        self.gcu_op = gcu_op
        self.topsflame_op = topsflame_op
        self.attribute = attribute
        self.op_type = None #从Attribute中获取的
        if attribute is None:
            self.op_type = gcu_op+" (None Attribute)"
        else:
            self.op_type = attribute.split("@")[0]
        self.input_shape = input_shape
        self.output_shape = output_shape
        self.input_size = input_size
        self.output_size = output_size
        self.total_size = total_size
        self.floats_1d = floats_1d
        self.floats_2d = floats_2d
        self.busy = busy
        if target is None:
            target = busy
        self.target = target
        self.gap = gap_to_pre
        self.others_gap = 0
        self.topsflame_gap = launch_kernel_gap + topsflame_duration
        if self.gap > self.topsflame_gap:
            self.others_gap = self.gap - self.topsflame_gap

        self.opt = max(busy - target, 0)  # ignore negtive opt
        # if topsflame_op == "DummyOp":
        #     self.opt = 0
        self.opt_pct = self.opt / busy
        self.band_width = total_size / busy
        self.Tflops = floats_2d / 1000 / busy
        if target > 0:
            self.band_width_targe = total_size / target
            self.Tflops_target = floats_2d / 1000 / target
        else:
            self.band_width_targe = float('nan')
            self.Tflops_target = float('nan')

        self.model_busy = -1
        self.busy_pct = -1
        self.opt_model_pct = -1

        self.top_topsflameop_handle = top_topsflameop_handle
        self.top_topsflameop_handle_name = top_topsflameop_handle_name
        self.flame_start = flame_start
        self.top_flame_start = top_flame_start
        self.hal_to_topsflameop_gap = int(flame_start) - int(top_flame_start)
        self.start_timestamp = start_timestamp
        self.end_timestamp = end_timestamp
        return

    def set_model_cost(self, model_busy):
        self.model_busy = model_busy
        self.busy_pct = self.busy / model_busy
        self.opt_model_pct = self.opt / model_busy
        return

    def __str__(self):
        return (f"index:{self.index:3d}    "
                f"cost:{self.busy:8d} ({self.busy_pct * 100:5.2f}%)    "
                f"gap:{self.gap:8d}    "
                f"target:{int(self.target):8d}    "
                f"opt: {self.opt_pct*100:4.1f}% (op) {self.opt_model_pct*100:4.1f}% (model)    "
                f"band width(GB/s): {self.band_width:5.1f} -> {self.band_width_targe:<5.1f}    "
                f"TFLOPS:{self.Tflops:5.2f} -> {self.Tflops_target:<5.2f}    "
                # f"floats_2d: {self.floats_2d}    "
                f"op_type: {self.op_type:<45}    "
                f"gcu_op: {self.gcu_op:<45}    "
                f"topsflame_op: {self.topsflame_op:<45}    "
                f"attribute: {self.attribute}")
    
    def gap_print(self):
        return (f"index:{self.index:3d}    "
                f"cost:{self.busy:8d} ({self.busy_pct * 100:5.2f}%)    "
                f"gap:{self.gap:8d}    "
                f"gap (others):{self.others_gap:8d}    "
                f"gap (topsflame):{self.topsflame_gap:8d}    "
                f"op_type: {self.op_type:<45}    "
                f"gcu_op: {self.gcu_op:<45}    "
                f"topsflame_op: {self.topsflame_op:<45}    "
                f"input_shape: {self.input_shape}                                                                                                                                                               "
                f"output_shape: {self.output_shape}    "
                f"attribute: {self.attribute}")


class ModelInfo:
    def __init__(self, model_name):
        self.model_name = model_name
        self.ops = [] # 1个model中一个step所有op相关信息
        self.op_types = {} # 1个model内的op分类统计（同一类型的数量、busy时间、优化时间、gap时间）
        self.cost = 0
        self.gap = 0
        self.others_gap = 0
        self.opt = 0
        self.opt_without_dummy = 0
        # self.target = 0
        self.dummy_cost = 0
        self.none_op_type_set = set()

        self.wx_target_attn_normalize_pattern = [
            "NormalizeOp", "NormalizeOp", "CopyTransposeOp", "MatmulOp"]
        self.wx_target_attn_normalize_matching_ops = []
        self.wx_target_attn_normalize_matched = []

        self.wx_mul_redcue_sum_pattern = ["BinaryOp", "SumOp"]
        self.wx_mul_redcue_sum_matching_ops = []
        self.wx_mul_redcue_sum_matched = []

        self.decomposed_fusions = []
        self.larg_gap_ops = []

    def InsertOP(self, op: GcuOPInfo):
        if op.op_type.find("(None Attribute)") > 0:
            if op.op_type not in self.none_op_type_set:
                self.none_op_type_set.add(op.op_type)
                print("[WARNING] empty attribute, gcu op: {}".format(op.gcu_op))

        self.ops.append(op)
        if op.op_type not in self.op_types:
            self.op_types[op.op_type] = [0, 0, 0, 0]
        self.op_types[op.op_type][0] += 1
        self.op_types[op.op_type][1] += op.busy
        self.op_types[op.op_type][2] += op.opt
        self.op_types[op.op_type][3] += op.gap
        self.cost += op.busy
        self.opt += op.opt
        self.gap += op.gap
        self.others_gap += op.others_gap
        # self.target += op.target
        if op.topsflame_op != "DummyOp":
            self.opt_without_dummy += op.opt
        else:
            self.dummy_cost += op.busy

        # # matching wx_target_attn_normalize
        # n = len(self.wx_target_attn_normalize_matching_ops)
        # if op.op_type == self.wx_target_attn_normalize_pattern[n]:
        #     self.wx_target_attn_normalize_matching_ops.append(op)
        # else:
        #     self.wx_target_attn_normalize_matching_ops = []
        # if len(self.wx_target_attn_normalize_matching_ops) == len(self.wx_target_attn_normalize_pattern):
        #     self.wx_target_attn_normalize_matched.append(
        #         self.wx_target_attn_normalize_matching_ops.copy())
        #     self.wx_target_attn_normalize_matching_ops = []

        # # matching wx_mul_redcue_sum
        # i = len(self.wx_mul_redcue_sum_matching_ops)
        # if op.op_type == self.wx_mul_redcue_sum_pattern[i] and ("Sum" in op.gcu_op or "MUL" in op.gcu_op):
        #     self.wx_mul_redcue_sum_matching_ops.append(op)
        # else:
        #     self.wx_mul_redcue_sum_matching_ops = []
        # if len(self.wx_mul_redcue_sum_matching_ops) == len(self.wx_mul_redcue_sum_pattern):
        #     self.wx_mul_redcue_sum_matched.append(
        #         self.wx_mul_redcue_sum_matching_ops.copy())
        #     self.wx_mul_redcue_sum_matching_ops = []
        # return

    def InsertDone(self):
        for op in self.ops:
            op.set_model_cost(self.cost)
        return

    def GetDecomposedFusions(self):
        class DecomposedFusion():
            def __init__(self, ops: list[GcuOPInfo]):
                self.fusion_ops = ops
                self.fusion_ops_top_name = ops[0].top_topsflameop_handle_name
                self.fusion_ops_top_handle = ops[0].top_topsflameop_handle
                self.kernels_cost = 0
            def GetKernelsCost(self):
                self.kernels_cost = int(self.fusion_ops[-1].end_timestamp) - int(self.fusion_ops[0].start_timestamp)

            def WriteDecomposedFusion(self, wb, sheet_name):

                if sheet_name not in wb.sheetnames:
                    sheet = wb.create_sheet(sheet_name)
                
                sheet = wb[sheet_name]
                column_titles = ["model", "top_topsflame_op_handle_name", "top_topsflame_op_handle", "op_type",
                    "kernel", "attribute", "input shapes", "output shapes", "busy time(ns)"]
                for i, title in enumerate(column_titles, start=1):
                    sheet.cell(1, i, title).font = Font(
                        bold=True, color=colors.BLACK)
                
                for op in self.fusion_ops:
                    row = len(tuple(sheet.rows)) + 1
                    sheet.cell(row, 1).value = op.model_name
                    sheet.cell(row, 2).value = self.fusion_ops_top_handle
                    sheet.cell(row, 3).value = self.fusion_ops_top_name
                    sheet.cell(row, 4).value = op.op_type
                    sheet.cell(row, 5).value = op.gcu_op
                    sheet.cell(row, 6).value = op.attribute
                    sheet.cell(row, 7).value = op.input_shape
                    sheet.cell(row, 8).value = op.output_shape
                    sheet.cell(row, 9).value = op.busy

        self.decomposed_fusions.clear()
        top_topsflame_handle_dict = {}
        for op_info in self.ops: # 要求 self.ops 一定要保持timeline 时间顺序！！！
            if op_info.top_topsflameop_handle not in top_topsflame_handle_dict:
                top_topsflame_handle_dict[op_info.top_topsflameop_handle] = [op_info]
            else:
                top_topsflame_handle_dict[op_info.top_topsflameop_handle].append(op_info)

        for k, v in top_topsflame_handle_dict.items():
            if len(v) > 1:
                decomposed_fusion = DecomposedFusion(v)
                self.decomposed_fusions.append(decomposed_fusion)
                decomposed_fusion.GetKernelsCost()

    def WriteDecomposedFusions(self, wb, sheet, kernels_cost_thresh=50000):
            for decomposed_fusion in self.decomposed_fusions:
                if decomposed_fusion.kernels_cost > kernels_cost_thresh:
                    decomposed_fusion.WriteDecomposedFusion(wb, "decomposed_fusions")

    def GetLargeGapOps(self, large_gap_thresh=5000, max_num=20):
        self.larg_gap_ops.clear()
        for op in self.ops:
            if op.gap >= large_gap_thresh:
                self.larg_gap_ops.append(op)
        self.larg_gap_ops.sort(key=lambda op: op.gap, reverse=True)
        if max_num > 0 and len(self.larg_gap_ops) > max_num:
            self.larg_gap_ops = self.larg_gap_ops[:max_num]
        return self.larg_gap_ops

    def WriteLargeGapOps(self, wb, sheet_name):
        if sheet_name not in wb.sheetnames:
            sheet = wb.create_sheet(sheet_name)
        
        sheet = wb[sheet_name]
        column_titles = ["model", "kernel", "busy time(ns)", "gap_to_pre_kernel", "gap1(hal_to_topsflameop)",
            "gap2(topsflameop_to_kernel)", "op_type", "input shapes", "output shapes", "attribute"]
        for i, title in enumerate(column_titles, start=1):
            sheet.cell(1, i, title).font = Font(
                bold=True, color=colors.BLACK)
        
        for op in self.larg_gap_ops:
            row = len(tuple(sheet.rows)) + 1
            colum = 1
            sheet.cell(row, colum).value = op.model_name
            colum+=1
            sheet.cell(row, colum).value = op.gcu_op
            colum+=1
            sheet.cell(row, colum).value = op.busy
            colum+=1
            sheet.cell(row, colum).value = op.gap
            colum+=1
            sheet.cell(row, colum).value = op.hal_to_topsflameop_gap
            colum+=1
            sheet.cell(row, colum).value = op.topsflame_gap
            colum+=1
            sheet.cell(row, colum).value = op.op_type
            colum+=1
            sheet.cell(row, colum).value = op.input_shape
            colum+=1
            sheet.cell(row, colum).value = op.output_shape
            colum+=1
            sheet.cell(row, colum).value = op.attribute

    def ShowOpTypes(self):
        def sort_func(e):
            return e['cost']
        print("info by op_type:")
        type_info_list = []
        for op_type, [num, cost, opt, gap] in self.op_types.items():
            type_info_list.append(
                {"op_type": op_type, "num": num, "cost": cost, "opt": opt, "gap": gap})
        type_info_list.sort(key=sort_func, reverse=True)
        for info in type_info_list:
            print("num:{:6d}    cost:{:5.1f}%    avg gap:{:8d}    opt:{:6.1f}% (op) {:4.1f}% (model)   type: {}".format(
                info['num'],
                info['cost'] / self.cost * 100,
                int(info['gap'] / info['num']),
                info['opt'] / info['cost'] * 100,
                info['opt'] / self.cost * 100,
                info['op_type'])
            )

    def ShowMatchedPattern(self):
        print("decomposed fusions:")
        total_cost = 0
        for matched in self.wx_target_attn_normalize_matched:
            for op in matched:
                total_cost += op.busy
        print("wx_target_attn_normalize num:{:3d}    total cost: {:.1f}%".format(
            len(self.wx_target_attn_normalize_matched), 100 * total_cost / self.cost))
        total_cost = 0
        for matched in self.wx_mul_redcue_sum_matched:
            for op in matched:
                total_cost += op.busy
        print("wx_mul_redcue_sum        num:{:3d}    total cost: {:.1f}%".format(
            len(self.wx_mul_redcue_sum_matched), 100 * total_cost / self.cost))

    def ShowLargeCostOp(self, cost_pct=0.01, cost=0):
        ops = []
        for op in self.ops:
            if op.busy_pct > cost_pct and op.busy > cost:
                ops.append(op)
        ops.sort(key=lambda op: op.cost, reverse=True)
        print("large cost op:")
        for op in ops:
            print(op)

    def ShowLargeOptOp(self, opt_pct=0.3):
        ops = []
        for op in self.ops:
            if op.opt_pct > opt_pct:
                ops.append(op)
        ops.sort(key=lambda op: op.opt/op.cost, reverse=True)
        print("large opt percentage op:")
        for op in ops:
            print(op)

    def ShowLargeOptOpOfModel(self):
        ops = []
        for op in self.ops:
            if op.opt_pct > 0.05 and op.opt >= 1000 and op.topsflame_op != "DummyOp":
                add = True
                for op_ in ops:
                    if op_.attribute == op.attribute:
                        add = False
                        break
                if add:
                    ops.append(op)
        ops.sort(key=lambda op: op.opt, reverse=True)
        # if len(ops) > 10:
        #     ops = ops[:10]
        print("opt ops:")
        for op in ops:
            print(op)

    def ShowLargeGapOpOfModel(self, threshold=5000, max_num=20):
        ops = []
        for op in self.ops:
            if op.gap >= threshold:
                ops.append(op)
        ops.sort(key=lambda op: op.gap, reverse=True)
        if max_num > 0 and len(ops) > max_num:
            ops = ops[:max_num]
        print("large gap ops:")
        for op in ops:
            print(op.gap_print())

    # 生成sheet, 每一行为每个可优化Op的相关信息
    def WriteOptOps(self, sheet):

        # 完全相同的Op合并统计
        class GcuOPMergeInfo():
            def __init__(self, op: GcuOPInfo):
                self.model_name = op.model_name
                self.op_type = op.op_type
                self.gcu_op = op.gcu_op
                self.attribute = op.attribute
                self.input_shape = op.input_shape
                self.output_shape = op.output_shape
                self.busy_min = self.busy_max = op.busy
                self.target = op.target
                self.gap = op.gap
                self.band_width_min = self.band_width_max = op.band_width
                self.band_width_target = op.band_width_targe
                self.Tflops_min = self.Tflops_max = op.Tflops
                self.Tflops_target = op.Tflops_target
                self.opt_model_pct = op.opt_model_pct
                self.num = 1

            def add(self, op: GcuOPInfo):
                self.busy_min = min(self.busy_min, op.busy)
                self.busy_max = max(self.busy_max, op.busy)
                self.band_width_min = min(self.band_width_min, op.band_width)
                self.band_width_max = max(self.band_width_max, op.band_width)
                self.Tflops_min = min(self.Tflops_min, op.Tflops)
                self.Tflops_max = max(self.Tflops_max, op.Tflops)
                self.opt_model_pct += op.opt_model_pct
                self.num += 1
                self.gap += op.gap

        ops = [] # 每个元素是GcuOPMergeInfo对象，记录n个相同Op的合并统计信息
        for op in self.ops:
            if op.opt_pct > 0.05 and op.opt >= 1000 and op.topsflame_op != "DummyOp": # 筛选出可优化的op
                new_op = True
                for op_ in ops:
                    if op_.attribute == op.attribute and op_.gcu_op == op.gcu_op:
                        new_op = False
                        op_.add(op)
                        break
                if new_op:
                    ops.append(GcuOPMergeInfo(op))
        # 按照对model的优化比例降序排列
        ops.sort(key=lambda op: op.opt_model_pct, reverse=True)

        # 逐行写入到表格，一行是num个op合并后的op相关信息
        for op in ops:
            row = len(tuple(sheet.rows)) + 1
            sheet.cell(row, 1).value = op.model_name
            sheet.cell(row, 2).value = op.op_type
            sheet.cell(row, 3).value = op.gcu_op
            sheet.cell(row, 4).value = op.attribute
            sheet.cell(row, 5).value = op.input_shape
            sheet.cell(row, 6).value = op.output_shape
            sheet.cell(row, 7).value = f"{op.busy_min}~{op.busy_max}"
            sheet.cell(row, 8).value = int(op.gap / op.num)
            sheet.cell(row, 9).value = int(op.target)
            sheet.cell(
                row, 10).value = f"[{op.band_width_min:.1f}~{op.band_width_max:.1f}] -> {op.band_width_target:.1f}"
            if op.Tflops_target > 0:
                sheet.cell(
                    row, 11).value = f"[{op.Tflops_min:.2f}~{op.Tflops_max:.2f}] -> {op.Tflops_target:.2f}"
            sheet.cell(row, 12).value = op.num
            sheet.cell(
                row, 13).value = f"{((op.busy_min - op.target) / op.busy_min * 100):.1f}%~{((op.busy_max - op.target) / op.busy_max * 100):.1f}%"
            sheet.cell(row, 14).value = op.opt_model_pct
            sheet.cell(row, 14).number_format = '0.00%'

    def printModelOptInfo(self):
        print("model name: {}".format(self.model_name))
        print("op numbers: {}".format(len(self.ops)))
        print("model cost(ns):  {:10d}".format(self.cost))
        print("model cost(ns):  {:10d}          -- with gaps ({:.1f}% are gaps and {:.1f}% from topsflame)".format(
            self.cost + self.gap, 100 * self.gap / (self.cost + self.gap),  100 * (self.gap - self.others_gap) / (self.cost + self.gap)))
        # print("model dummy cost: {0:8d}".format(self.dummy_cost))
        print("model opt(ns):   {:10d} ({:4.1f}%)".format(
            int(self.opt), self.opt / self.cost * 100))
        print("                 {:10d} ({:4.1f}%)  -- without dummy op".format(
            int(self.opt_without_dummy), self.opt_without_dummy / self.cost * 100))
        print("model target(ns):{:10d}".format(int(self.cost - self.opt)))
        print("                 {:10d}          -- without dummy op".format(
            int(self.cost - self.opt_without_dummy)))

    # 插入表格一行，记录该model可优化情况
    def WriteModelSummaryInfo(self, sheet):
        row = len(tuple(sheet.rows)) + 1
        sheet.cell(row, 1).value = self.model_name
        sheet.cell(row, 2).value = len(self.ops)
        sheet.cell(row, 3).value = self.cost
        sheet.cell(row, 4).value = self.gap
        sheet.cell(row, 5).value = self.gap - self.others_gap
        sheet.cell(row, 6).value = self.others_gap
        sheet.cell(row, 7).value = int(self.opt_without_dummy)
        sheet.cell(row, 8).value = self.opt_without_dummy / self.cost
        sheet.cell(row, 8).number_format = '0.00%'
        sheet.cell(row, 9).value = int(self.cost - self.opt_without_dummy)


class WholeInfo():
    key_map = {
        "index": 1,
        "start_timestamp": 2,
        "end_timestamp": 3,
        "op_id": 4,
        "gcu_op": 5,
        "topsflame_op": 6,
        "input_shape": 7,
        "output_shape": 8,
        "attribute": 9,
        "floats_1d": 10,
        "1d_sfu": 11,
        "floats_2d": 12,
        "input_size": 13,
        "output_size": 14,
        "total_size": 15,
        "SFU_floats": 16,
        "1D_flops_ratio": 17,
        "2D_flops_ratio": 18,
        "SFU_flops_ratio": 19,
        "busy": 20,
        "gap": 21,
        "topsflameop_handle": 22,
        "top_topsflameop_handle": 23,
        "top_topsflameop_handle_name": 24,
        "top_topsflameop_start_timestamp" : 25,
        "top_topsflameop_end_timestamp" : 26,
        "topsflameop_start_timestamp": 27,
        "topsflameop_end_timestamp": 28,
        "topsflameop_duration": 29,
        "launch_kernel_to_op_run_gap": 30,
        "runtime_start_timestamp": 31,
        "runtime_end_timestamp": 32,
        "runtime_duration_timestamp": 33,
        "target": 34,
        "bound": 35,
        "device_id": 36,
        "stream_id": 37,
        "host_name": 38,
        "step": 39,
    }

    # 从表格中读出目标行的op数据，生产GcuOPInfo对象
    def getOP(self, sheet, row: int, model_name: str):
        op_id = sheet.cell(row, self.key_map['op_id']).value
        index = sheet.cell(row, self.key_map['index']).value
        gcu_op = sheet.cell(row, self.key_map['gcu_op']).value
        topsflame_op = sheet.cell(row, self.key_map['topsflame_op']).value
        attribute = sheet.cell(row, self.key_map['attribute']).value
        input_shape = sheet.cell(row, self.key_map['input_shape']).value
        output_shape = sheet.cell(row, self.key_map['output_shape']).value
        input_size = sheet.cell(row, self.key_map['input_size']).value
        output_size = sheet.cell(row, self.key_map['output_size']).value
        total_size = sheet.cell(row, self.key_map['total_size']).value
        floats_1d = sheet.cell(row, self.key_map['output_size']).value
        floats_2d = sheet.cell(row, self.key_map['floats_2d']).value
        busy = int(sheet.cell(row, self.key_map['busy']).value)
        target = sheet.cell(row, self.key_map['target']).value
        gap = sheet.cell(row, self.key_map['gap']).value
        start_timestamp = sheet.cell(row, self.key_map['start_timestamp']).value
        end_timestamp = sheet.cell(row, self.key_map['end_timestamp']).value

        flame_start = sheet.cell(row, self.key_map['topsflameop_start_timestamp']).value
        flame_end = sheet.cell(row, self.key_map['topsflameop_end_timestamp']).value
        top_flame_start = sheet.cell(row, self.key_map['top_topsflameop_start_timestamp']).value
        top_flame_end = sheet.cell(row, self.key_map['top_topsflameop_end_timestamp']).value
        topsflame_duration = int(sheet.cell(row, self.key_map['topsflameop_duration']).value)
        launch_kernel_gap = int(sheet.cell(row, self.key_map['launch_kernel_to_op_run_gap']).value)
        top_topsflameop_handle = sheet.cell(row, self.key_map['top_topsflameop_handle']).value
        top_topsflameop_handle_name = sheet.cell(row, self.key_map['top_topsflameop_handle_name']).value



        op = GcuOPInfo(model_name, op_id, index, gcu_op, topsflame_op, attribute,
                       input_shape, output_shape, input_size, output_size, total_size,
                       floats_1d, floats_2d, busy, target, gap, flame_start, top_flame_start, topsflame_duration, launch_kernel_gap, 
                       top_topsflameop_handle, top_topsflameop_handle_name, start_timestamp, end_timestamp)
        return op

    def __init__(self, file_list):
        self.models_info_list = []
        self.models_info_dict = {}
        self.op_types = {} #全部模型按op类别统计（num、cost、opt、gap)
        self.cost = 0
        self.opt = 0
        self.opt_without_dummy = 0
        opt_ops_ = {}

        print("start to load reports.")
        for filename in file_list:
            wb = load_workbook(filename=filename)
            model_name = filename.split("/")[-1][:-5]
            print(model_name, flush=True)
            sheet = wb["analytics_sheet"]
            row_size = len(tuple(sheet.rows))
            steps = sheet.cell(row_size, self.key_map['step']).value
            print(f"[INFO] profile {model_name} have {steps} step.")

            # 设置要获取的目标step，以及目标step数据的起始行和结束行
            start_row = 2
            end_row = row_size
            if (row_size - 1) % steps == 0:
                model_op_num = int((row_size - 1) / steps)
                print(f"each step have {model_op_num} ops.")
                select_step = steps - 1 # 取倒数第二个step数据（如果steps大于1）
                if select_step <= 1:
                    select_step = steps
                start_row = (select_step - 1) * model_op_num + 2
                end_row = start_row + model_op_num - 1
                print(f"select {select_step} step ops.")
            else:
                print(f"total have {row_size - 1} ops.")

            # 从表格中获取并生成一个model的一个目标step全部数据的对象；并保存可优化的op到opt_ops_字典.
            model_info = ModelInfo(model_name)
            for i in range(start_row, end_row + 1):
                op = self.getOP(sheet, i, model_name)
                if i == start_row:
                    op.gap = 0
                    op.others_gap = 0
                model_info.InsertOP(op)
                if op.topsflame_op != "DummyOp" and op.opt > 0 and op.attribute not in opt_ops_:
                    opt_ops_[op.attribute] = op
            model_info.InsertDone()

            # 将单个model_info对象添加到总的list和dict
            self.models_info_list.append(model_info)
            self.models_info_dict[model_name] = model_info

            # 所有model按op类别统计信息保存在self.op_types字典中
            for op_type, [num, cost, opt, gap] in model_info.op_types.items():
                if op_type not in self.op_types:
                    self.op_types[op_type] = [0, 0, 0, 0]
                self.op_types[op_type][0] += num
                self.op_types[op_type][1] += cost
                self.op_types[op_type][2] += opt
                self.op_types[op_type][3] += gap

        # 按照模型名字排序
        self.models_info_list.sort(key=lambda m: m.model_name)

        self.opt_ops = []
        for k, v in opt_ops_.items():
            self.opt_ops.append(v)

        self.GetAllModelsDecomposedFusions()
        self.GetAllModelsLargeGapOps()
        print("load reports done.")
    
    # 找出topsflame顶层是一个op，但下面lauch多个op的情况
    def GetAllModelsDecomposedFusions(self):
        for model_info in self.models_info_list:
            model_info.GetDecomposedFusions()
    
    def GetAllModelsLargeGapOps(self):
        self.all_models_large_gap_ops_dict = {}
        for model_info in self.models_info_list:
            large_gap_ops = model_info.GetLargeGapOps()
            self.all_models_large_gap_ops_dict[model_info.model_name] = large_gap_ops
        return self.all_models_large_gap_ops_dict

    def ShowAllModelsDecomposedFusions(self):
        for k, v in self.all_decomposed_fusions_dict.items():
            print(k)
            for fusion in v:
                print(fusion.fusion_ops_top_name)

    def ShowBiasInfo(self):
        print("basic info:")
        print("model num:   {:4d}".format(len(self.models_info_list)))
        print("op_type num: {:4d}".format(len(self.op_types)))

    def ShowTypesInfo(self):
        type_info_list = []
        cost_sum = 0
        for op_type, [num, cost, opt, gap] in self.op_types.items():
            type_info_list.append(
                {'op_type': op_type, 'num': num, 'cost': cost, 'opt': opt, 'gap': gap})
            cost_sum += cost
        type_info_list.sort(key=lambda i: i['cost'], reverse=True)

        print("info by op_type of all models:")
        for info in type_info_list:
            print("num:{:6d}    cost:{:5.1f}%    avg gap:{:8d}    opt:{:6.1f}% (op) {:4.1f}% (model)   type: {}".format(
                info['num'],
                info['cost'] / cost_sum * 100,
                int(info['gap'] / info['num']),
                info['opt'] / info['cost'] * 100,
                info['opt'] / cost_sum * 100,
                info['op_type'])
            )
        return

    def ShowOPsInfo(self):
        self.opt_ops.sort(key=lambda op: (op.op_type, -op.opt), reverse=False)
        print("info by op attibute of all models:")
        for op in self.opt_ops:
            print(op)

    def ShowModelInfo(self, model_names=[]):
        print("info of each model:")
        if model_names is None or len(model_names) == 0:
            for model_info in self.models_info_list:
                print("-----"*20)
                model_info.printModelOptInfo()
                print("")
                model_info.ShowOpTypes()
                print("")
                model_info.ShowMatchedPattern()
                print("")
                model_info.ShowLargeOptOpOfModel()
                print("")
                model_info.ShowLargeGapOpOfModel()
        else:
            for model_name in model_names:
                if model_name in self.models_info_dict:
                    print("-----"*20)
                    model_info = self.models_info_dict[model_name]
                    model_info.printModelOptInfo()
                    print("")
                    model_info.ShowOpTypes()
                    print("")
                    model_info.ShowMatchedPattern()
                    print("")
                    model_info.ShowLargeOptOpOfModel()
                    print("")
                    model_info.ShowLargeGapOpOfModel()

    # 生成最终后处理的表格，有两个sheet，分别是每个model的可优化情况和每个op的可优化情况
    def SaveModelInfo(self, save_name):
        wb = Workbook()
        del wb[wb.active.title]
        models_sheet = wb.create_sheet("models-summary")
        column_titles = ["model", "op num",
                         "cost(ns)", "gaps(ns)", "topsflame gaps(ns)", "others gaps(ns)", "opt(ns)", "opt(%)", "target(ns)"]
        for i, title in enumerate(column_titles, start=1):
            models_sheet.cell(1, i, title).font = Font(
                bold=True, color=colors.BLACK)

        ops_sheet = wb.create_sheet("ops")
        column_titles = ["model", "op_type", "kernel", "attribute", "input shapes",
                         "output shapes", "busy time(ns)", "avg gap(ns)", "target(ns)", "BW(GB/s)",
                         "Tflops", "num", "opt(of op)", "opt(of model)"]
        for i, title in enumerate(column_titles, start=1):
            ops_sheet.cell(1, i, title).font = Font(
                bold=True, color=colors.BLACK)

        for model_info in self.models_info_list:
            model_info.WriteModelSummaryInfo(models_sheet)

            # 一次写一个模型的op信息到ops_sheet
            model_info.WriteOptOps(ops_sheet)

        for model_info in self.models_info_list:
            model_info.WriteDecomposedFusions(wb, "decomposed_fusions")
        
        for model_info in self.models_info_list:
            model_info.WriteLargeGapOps(wb, "large_gap_ops")

        wb.save(save_name)
        print("[INFO] save reslult file:", save_name)


def find_all_excel(path):
    file_list = []
    with os.scandir(path) as entries:
        for entry in entries:
            if entry.is_file():
                if entry.name[-5:] == ".xlsx":
                    file_list.append(entry.path)
    return file_list


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    # 所有excel文件需要直接存放在该目录内, 不能放在子目录内, excel文件名建议为model_name.xlsx
    parser.add_argument("--report", type=str,
                        help='folder where TopsAnalytics report store in')
    parser.add_argument("--save-name", type=str, default=None,
                        help='file to save ops info')
    args = parser.parse_args()
    if args.save_name is not None:
        save_name = args.save_name
        if save_name[-5:] != ".xlsx":
            save_name += ".xlsx"
    else:
        save_name = args.report
        if save_name[-1] == '/':
            save_name = save_name[:-1]
        save_name += "_info.xlsx"

    file_list = find_all_excel(args.report)
    print("====" * 25)
    whole_info = WholeInfo(file_list)
    print("====" * 25)
    whole_info.ShowBiasInfo()
    print("====" * 25)
    whole_info.ShowTypesInfo()
    # print("====" * 25)
    # whole_info.ShowOPsInfo()
    print("====" * 25)
    whole_info.ShowModelInfo()
    print("====" * 25)
    whole_info.SaveModelInfo(save_name)
